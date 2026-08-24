"""
Convert a filled machine-registry workbook into YAML records.

    python scripts/import_from_xlsx.py intake/01_machine_registry.xlsx

Reads the "Autonomous Registry" and "Manual Registry" sheets, merges each row into the
matching file in data/machines/ (matched on vendor + model), and reports what changed.

Existing values are only overwritten when the workbook supplies something and the YAML has
null. That way a hand-curated note or a measured coefficient is never clobbered by a blank
cell. Pass --overwrite to force.
"""
import argparse
import glob
import os
import sys

import yaml

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl required:  pip install -r requirements.txt")

MACHINE_DIR = "data/machines"

# workbook column header -> dotted path in the YAML record
FIELD_MAP = {
    "cut_width_in": "cutting.width_in",
    "hoc_min_in": "cutting.hoc_min_in",
    "hoc_max_in": "cutting.hoc_max_in",
    "battery_kwh": "power.battery_kwh",
    "runtime_hr": "power.runtime_hr",
    "charge_hr": "power.charge_hr",
    "fast_charge_hr": "power.fast_charge_hr",
    "fuel_gal_per_hr": "power.fuel_gal_per_hr",
    "ground_speed_mph": "capacity.ground_speed_mph",
    "machine_life_hr": "mechanical.machine_life_hr",
    "service_interval_hr": "mechanical.service_interval_hr",
    "annual_service_usd": "mechanical.annual_service_usd",
    "noise_db": "mechanical.noise_db",
    "weight_lb": "mechanical.weight_lb",
    "list_price_usd": "cost.list_price_usd",
    "infra_cost_usd": "cost.infrastructure_usd",
    "annual_subscription_usd": "cost.annual_subscription_usd",
}


def set_path(obj, dotted, value):
    keys = dotted.split(".")
    for k in keys[:-1]:
        obj = obj.setdefault(k, {})
    obj[keys[-1]] = value


def get_path(obj, dotted):
    for k in dotted.split("."):
        if not isinstance(obj, dict):
            return None
        obj = obj.get(k)
    return obj


def load_records():
    out = {}
    for path in glob.glob(f"{MACHINE_DIR}/*.yaml"):
        if os.path.basename(path).startswith("_"):
            continue
        rec = yaml.safe_load(open(path))
        key = (str(rec.get("vendor", "")).strip().lower(),
               str(rec.get("model", "")).strip().lower())
        out[key] = (path, rec)
    return out


def sheet_rows(ws):
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        yield dict(zip(headers, row))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--overwrite", action="store_true",
                    help="replace existing values, not just fill nulls")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    wb = load_workbook(args.workbook, data_only=True)
    records = load_records()
    changes, unmatched = [], []

    for sheet in ("Autonomous Registry", "Manual Registry"):
        if sheet not in wb.sheetnames:
            continue
        for row in sheet_rows(wb[sheet]):
            key = (str(row.get("vendor") or "").strip().lower(),
                   str(row.get("model") or "").strip().lower())
            if key not in records:
                if key != ("", ""):
                    unmatched.append(f"{row.get('vendor')} {row.get('model')}")
                continue
            path, rec = records[key]
            for col, dotted in FIELD_MAP.items():
                new = row.get(col)
                if new is None or new == "":
                    continue
                if dotted == "cost.list_price_usd":
                    cur = get_path(rec, "cost.currency_original")
                    if cur and cur != "USD":
                        print(f"  SKIP {os.path.basename(path)} list_price_usd: record is "
                              f"priced in {cur}. Convert to USD and set currency_original "
                              f"before importing.")
                        continue
                old = get_path(rec, dotted)
                if old is not None and not args.overwrite:
                    continue
                if old == new:
                    continue
                set_path(rec, dotted, new)
                changes.append(f"{os.path.basename(path)}  {dotted}: {old!r} -> {new!r}")
            if not args.dry_run:
                with open(path, "w") as f:
                    yaml.dump(rec, f, sort_keys=False, default_flow_style=False,
                              allow_unicode=True, width=100)

    print(f"{len(changes)} field(s) updated"
          + (" (dry run, nothing written)" if args.dry_run else ""))
    for c in changes:
        print("  " + c)
    if unmatched:
        print(f"\n{len(unmatched)} workbook row(s) had no matching YAML file:")
        for u in unmatched:
            print("  " + u)
        print("  Create a record first: copy data/machines/_TEMPLATE.yaml")
    print("\nNext:  python scripts/validate.py")


if __name__ == "__main__":
    main()
